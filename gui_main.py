from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QLineEdit,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QFileDialog,
    QMessageBox,
    QComboBox,
    QListWidget,
    QAbstractItemView,
    QProgressDialog,
)
from PyQt5.QtCore import Qt, QObject, QThread, QTimer, pyqtSignal, pyqtSlot
import json
import os
import pandas as pd
from db_utils import get_connection, get_tables, get_table_schema, get_table_ddl, export_schema_to_excel

SETTINGS_FILE = 'c:/Python/TABLE_SET/settings.json'


class ExcelExportWorker(QObject):
    progress = pyqtSignal(int, str)  # (done_tables, message)
    finished = pyqtSignal(str)  # save_path
    error = pyqtSignal(str)

    def __init__(
        self,
        dbtype: str,
        host: str,
        port: int,
        user: str,
        pw: str,
        db: str,
        tables: list[str],
        table_comments: dict[str, str],
        save_path: str,
    ):
        super().__init__()
        self._dbtype = dbtype
        self._host = host
        self._port = port
        self._user = user
        self._pw = pw
        self._db = db
        self._tables = tables
        self._table_comments = table_comments
        self._save_path = save_path

    @pyqtSlot()
    def run(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            # NOTE: DB 커넥션은 스레드 내부에서 새로 생성 (스레드 안전)
            conn = get_connection(
                self._dbtype,
                self._host,
                self._port,
                self._user,
                self._pw,
                self._db,
            )

            wb = openpyxl.Workbook()

            # 1. 테이블 목록 시트
            ws1 = wb.active
            ws1.title = '테이블 목록'
            ws1.append(['NN', 'Table Name', 'Description'])
            for idx, t in enumerate(self._tables, 1):
                ws1.append([idx, t, self._table_comments.get(t, '')])

            header_fill = PatternFill('solid', fgColor='B7B7B7')
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(1, 4):
                ws1.column_dimensions[get_column_letter(col)].width = 20

            # 2. 테이블 정의서 시트
            ws2 = wb.create_sheet('테이블 정의서')
            row_cursor = 1
            table_blocks: list[tuple[int, int]] = []

            for done, t in enumerate(self._tables, 1):
                self.progress.emit(done - 1, f'스키마 조회 중: {t}')
                block_start = row_cursor
                schema = get_table_schema(conn, self._dbtype, t)
                columns = schema['columns']
                pk = schema['primary_key']
                fk = schema['foreign_keys']
                indexes = schema['indexes']
                comment = schema['table_comment']

                # 테이블 정보 블록
                ws2.merge_cells(start_row=row_cursor, start_column=2, end_row=row_cursor, end_column=8)
                ws2.cell(row=row_cursor, column=2, value='테이블 정의서').fill = header_fill
                ws2.cell(row=row_cursor, column=2).font = Font(bold=True)
                ws2.cell(row=row_cursor, column=2).alignment = Alignment(horizontal='center')
                row_cursor += 1

                ws2.cell(row=row_cursor, column=2, value='Table Name')
                ws2.cell(row=row_cursor, column=3, value=t)
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Table ID')
                ws2.cell(row=row_cursor, column=3, value=t)
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Description')
                ws2.cell(row=row_cursor, column=3, value=comment)
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Primary Key')
                ws2.cell(row=row_cursor, column=3, value=','.join(pk) if pk else '')
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Foreign Key')
                fk_str = ', '.join([f"{c}->{t2}({col})" for c, t2, col in fk]) if fk else ''
                ws2.cell(row=row_cursor, column=3, value=fk_str)
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Index info #1')
                row_cursor += 1

                if indexes:
                    idx_headers = ['Index Name', 'Columns', 'Unique']
                    for i, h in enumerate(idx_headers, 2):
                        ws2.cell(row=row_cursor, column=i, value=h)
                        ws2.cell(row=row_cursor, column=i).fill = header_fill
                        ws2.cell(row=row_cursor, column=i).font = Font(bold=True)
                        ws2.cell(row=row_cursor, column=i).alignment = Alignment(horizontal='center')
                    row_cursor += 1
                    for iname, icols, unique in indexes:
                        ws2.cell(row=row_cursor, column=2, value=iname)
                        ws2.cell(row=row_cursor, column=3, value=icols)
                        ws2.cell(row=row_cursor, column=4, value=unique)
                        row_cursor += 1
                else:
                    row_cursor += 1

                col_headers = ['NN', 'Physical Name', 'Logical Name', 'Data Type', 'Null', 'Key', 'Default']
                for i, h in enumerate(col_headers, 2):
                    ws2.cell(row=row_cursor, column=i, value=h)
                    ws2.cell(row=row_cursor, column=i).fill = header_fill
                    ws2.cell(row=row_cursor, column=i).font = Font(bold=True)
                    ws2.cell(row=row_cursor, column=i).alignment = Alignment(horizontal='center')
                row_cursor += 1

                # 컬럼 정보 (기존 구현 그대로: MySQL 컬럼 튜플 포맷에 최적화)
                for idx, col in enumerate(columns, 1):
                    null_val = 'NN' if len(col) > 3 and col[3] == 'NO' else ''
                    key_raw = col[4] if len(col) > 4 else ''
                    if key_raw == 'PRI':
                        key_val = 'PK'
                    elif key_raw == 'MUL':
                        key_val = 'MUL'
                    else:
                        key_val = key_raw
                    ws2.cell(row=row_cursor, column=2, value=idx)
                    ws2.cell(row=row_cursor, column=3, value=col[0] if len(col) > 0 else '')
                    ws2.cell(row=row_cursor, column=4, value=(col[8] if len(col) > 8 else ''))
                    ws2.cell(row=row_cursor, column=5, value=col[1] if len(col) > 1 else '')
                    ws2.cell(row=row_cursor, column=6, value=null_val)
                    ws2.cell(row=row_cursor, column=7, value=key_val)
                    ws2.cell(row=row_cursor, column=8, value=(col[5] if len(col) > 5 else ''))
                    row_cursor += 1

                block_end = row_cursor - 1
                table_blocks.append((block_start, block_end))
                row_cursor += 2

                self.progress.emit(done, f'엑셀 작성 중: {t}')

            # 스타일: 전체 표 테두리(테이블 정의서 시트는 각 테이블 블록에만 적용)
            thin = Side(border_style='thin', color='000000')
            for start, end in table_blocks:
                for row in ws2.iter_rows(min_row=start, max_row=end, min_col=2, max_col=8):
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for row in ws1.iter_rows():
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # 자동 높이/너비
            for ws in [ws1, ws2]:
                for col in ws.columns:
                    maxlen = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                maxlen = max(maxlen, len(str(cell.value)))
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = max(12, min(maxlen + 2, 40))

            self.progress.emit(len(self._tables), '파일 저장 중...')
            wb.save(self._save_path)
            try:
                conn.close()
            except Exception:
                pass

            self.finished.emit(self._save_path)
        except PermissionError:
            self.error.emit(
                f'파일이 열려 있어 저장할 수 없습니다.\n파일을 닫고 다시 시도하세요:\n{self._save_path}'
            )
        except Exception as exc:
            self.error.emit(f'엑셀 저장 중 오류 발생: {exc}')


class SchemaLoadWorker(QObject):
    finished = pyqtSignal(int, str, list, list)  # request_id, info_html, headers, rows
    error = pyqtSignal(int, str)

    def __init__(
        self,
        request_id: int,
        dbtype: str,
        host: str,
        port: int,
        user: str,
        pw: str,
        db: str,
        tables: list[str],
    ):
        super().__init__()
        self._request_id = request_id
        self._dbtype = dbtype
        self._host = host
        self._port = port
        self._user = user
        self._pw = pw
        self._db = db
        self._tables = tables

    @pyqtSlot()
    def run(self):
        try:
            conn = get_connection(
                self._dbtype,
                self._host,
                self._port,
                self._user,
                self._pw,
                self._db,
            )

            info_html = ''
            all_rows: list[list] = []
            all_headers = None

            for table in self._tables:
                schema = get_table_schema(conn, self._dbtype, table)
                columns = schema['columns']
                indexes = schema['indexes']
                pk = schema['primary_key']
                fk = schema['foreign_keys']
                comment = schema['table_comment']

                info = f"<b>Table:</b> {table}<br>"
                info += f"<b>Comment:</b> {comment}<br>"
                info += f"<b>Primary Key:</b> {', '.join(pk) if pk else '-'}<br>"
                if fk:
                    fk_str = '<br>'.join([f"{c} → {t}({col})" for c, t, col in fk])
                else:
                    fk_str = '-'
                info += f"<b>Foreign Key:</b> {fk_str}<br>"

                if indexes:
                    # MySQL: (index_name, columns, unique)
                    idx_str = '<br>'.join(
                        [f"<b>{row[0]}</b> <span style='color:blue'>{row[1]}</span> ({row[2]})" for row in indexes]
                    )
                else:
                    idx_str = '-'
                info += f"<b>Index info:</b> {idx_str}"
                info_html += info + '<hr>'

                if self._dbtype == 'mysql':
                    headers = ['No', 'Table', 'Field', 'Type', 'Null', 'Key', 'Default', 'Extra', 'Comment']
                    if all_headers is None:
                        all_headers = headers
                    for idx, col in enumerate(columns, 1):
                        null_val = 'NN' if len(col) > 3 and col[3] == 'NO' else ''
                        row = [
                            idx,
                            table,
                            col[0] if len(col) > 0 else '',
                            col[1] if len(col) > 1 else '',
                            null_val,
                            col[4] if len(col) > 4 else '',
                            col[5] if len(col) > 5 else '',
                            col[6] if len(col) > 6 else '',
                            col[8] if len(col) > 8 else '',
                        ]
                        all_rows.append(row)
                    all_rows.append([''] * len(headers))
                elif self._dbtype == 'postgresql':
                    headers = ['No', 'Table', 'column_name', 'data_type', 'is_nullable', 'column_default']
                    if all_headers is None:
                        all_headers = headers
                    for idx, col in enumerate(columns, 1):
                        row = [idx, table] + list(col)
                        all_rows.append(row)
                    all_rows.append([''] * len(headers))
                else:
                    # 기타 DB는 기존 UI가 구조 표를 특정 포맷으로 가정하고 있어
                    # 테이블 정보만 표시하고 구조 표는 비워둠(기존 동작과 동일하게 최소 변경)
                    if all_headers is None:
                        all_headers = []

            if all_headers and all_rows and all_rows[-1] == [''] * len(all_headers):
                all_rows.pop()

            try:
                conn.close()
            except Exception:
                pass

            self.finished.emit(self._request_id, info_html, all_headers or [], all_rows)
        except Exception as exc:
            self.error.emit(self._request_id, str(exc))

class MainWindow(QWidget):
    def handle_table_cell_click(self, row, col):
        # 테이블명 셀 클릭 시 체크박스 토글
        if col == 1:
            item = self.table_list.item(row, 0)
            if item:
                item.setCheckState(0 if item.checkState() else 2)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("DB Table Viewer (확장판)")
        self.conn = None
        self.dbtype = 'mysql'
        self._conn_params = None
        self._export_thread = None
        self._export_worker = None
        self._export_progress = None
        self._schema_timer = QTimer(self)
        self._schema_timer.setSingleShot(True)
        self._schema_timer.timeout.connect(self._start_schema_reload)
        self._schema_request_id = 0
        self._schema_thread = None
        self._schema_worker = None

        # 대량 테이블 구조 표시를 청크로 렌더링
        self._populate_timer = QTimer(self)
        self._populate_timer.setSingleShot(False)
        self._populate_timer.timeout.connect(self._populate_next_chunk)
        self._populate_request_id = 0
        self._populate_rows = []
        self._populate_index = 0
        self._populate_chunk_size = 200

        self.init_ui()
        self.load_settings()

    def init_ui(self):
        layout = QVBoxLayout()
        # DB 입력 폼
        form_layout = QHBoxLayout()
        self.dbtype_combo = QComboBox(); self.dbtype_combo.addItems(['mysql', 'postgresql', 'cubrid'])
        self.host_input = QLineEdit(); self.host_input.setPlaceholderText('Host')
        self.port_input = QLineEdit(); self.port_input.setPlaceholderText('Port')
        self.user_input = QLineEdit(); self.user_input.setPlaceholderText('User')
        self.pw_input = QLineEdit(); self.pw_input.setPlaceholderText('Password'); self.pw_input.setEchoMode(QLineEdit.Password)
        self.db_input = QLineEdit(); self.db_input.setPlaceholderText('Database')
        self.connect_btn = QPushButton('Connect'); self.connect_btn.clicked.connect(self.connect_db)
        self.pw_input.returnPressed.connect(self.connect_db)
        form_layout.addWidget(QLabel('DB Type'))
        form_layout.addWidget(self.dbtype_combo)
        form_layout.addWidget(self.host_input)
        form_layout.addWidget(self.port_input)
        form_layout.addWidget(self.user_input)
        form_layout.addWidget(self.pw_input)
        form_layout.addWidget(self.db_input)
        form_layout.addWidget(self.connect_btn)
        layout.addLayout(form_layout)

        # 테이블 목록 (체크박스/전체선택)
        from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QCheckBox, QHeaderView
        self.select_all_checkbox = QCheckBox('전체선택')
        self.select_all_checkbox.stateChanged.connect(self.select_all_tables)
        layout.addWidget(self.select_all_checkbox)
        self.table_list = QTableWidget()
        self.table_list.setColumnCount(2)
        self.table_list.setHorizontalHeaderLabels(['선택', '테이블 (이름 / 코멘트)'])
        self.table_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table_list.cellChanged.connect(self.show_table_schema)
        self.table_list.cellClicked.connect(self.handle_table_cell_click)
        layout.addWidget(self.table_list)

        # 테이블 정보(이름, 코멘트, PK, FK, 인덱스) 표시 (스크롤 가능)
        from PyQt5.QtWidgets import QScrollArea
        self.info_label = QLabel('')
        self.info_label.setTextInteractionFlags(self.info_label.textInteractionFlags() | Qt.TextSelectableByMouse)
        self.info_label.setWordWrap(True)
        self.info_scroll = QScrollArea()
        self.info_scroll.setWidgetResizable(True)
        self.info_scroll.setWidget(self.info_label)
        self.info_scroll.setMinimumHeight(180)
        layout.addWidget(self.info_scroll)

        # 테이블 구조 표시
        self.schema_table = QTableWidget()
        layout.addWidget(QLabel('테이블 구조'))
        layout.addWidget(self.schema_table)
        self.schema_table.verticalHeader().setVisible(False)

        # 버튼들
        btn_layout = QHBoxLayout()
        self.export_btn = QPushButton('엑셀로 내보내기'); self.export_btn.clicked.connect(self.export_to_excel)
        self.ddl_btn = QPushButton('DDL 생성'); self.ddl_btn.clicked.connect(self.show_ddl)
        self.save_btn = QPushButton('접속정보 저장'); self.save_btn.clicked.connect(self.save_settings)
        btn_layout.addWidget(self.export_btn)
        btn_layout.addWidget(self.ddl_btn)
        btn_layout.addWidget(self.save_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def connect_db(self):
        dbtype = self.dbtype_combo.currentText()
        host = self.host_input.text()
        port = int(self.port_input.text())
        user = self.user_input.text()
        pw = self.pw_input.text()
        db = self.db_input.text()
        try:
            self.conn = get_connection(dbtype, host, port, user, pw, db)
            self.dbtype = dbtype
            tables = get_tables(self.conn, dbtype)
            self._conn_params = (dbtype, host, port, user, pw, db)

            # 대량 테이블 로딩 시 시그널 폭주/리페인트 비용 감소
            self.table_list.setUpdatesEnabled(False)
            self.table_list.blockSignals(True)
            self.table_list.setRowCount(len(tables))
            for i, (name, comment) in enumerate(tables):
                display = f"{name} / {comment}" if comment else name
                chk_item = QTableWidgetItem()
                chk_item.setCheckState(0)
                self.table_list.setItem(i, 0, chk_item)
                self.table_list.setItem(i, 1, QTableWidgetItem(display))
            self.table_list.blockSignals(False)
            self.table_list.setUpdatesEnabled(True)
            self.save_settings()  # 연결 성공 시 자동 저장
            QMessageBox.information(self, '성공', 'DB 연결 및 테이블 목록 불러오기 성공!')
        except Exception as e:
            QMessageBox.critical(self, '오류', str(e))
    def select_all_tables(self, state):
        # 전체선택 시 cellChanged가 행 개수만큼 발생하면 UI가 멈출 수 있어 blockSignals
        self.table_list.blockSignals(True)
        for i in range(self.table_list.rowCount()):
            item = self.table_list.item(i, 0)
            if item:
                item.setCheckState(2 if state else 0)
        self.table_list.blockSignals(False)
        self.show_table_schema()

    def show_table_schema(self, *_args):
        # 디바운스: 연속 체크/변경 시 과도한 재조회 방지
        if not self.conn or not self._conn_params:
            return
        self._schema_timer.start(150)

    def _start_schema_reload(self):
        if not self.conn or not self._conn_params:
            return

        # 이전 렌더링 작업 중단
        if self._populate_timer.isActive():
            self._populate_timer.stop()
        self._populate_rows = []
        self._populate_index = 0

        checked_tables: list[str] = []
        for i in range(self.table_list.rowCount()):
            item = self.table_list.item(i, 0)
            if item and item.checkState():
                display = self.table_list.item(i, 1).text()
                checked_tables.append(display.split(' / ')[0])

        if not checked_tables:
            self.info_label.setText('')
            self.schema_table.clear()
            return

        self._schema_request_id += 1
        request_id = self._schema_request_id

        dbtype, host, port, user, pw, db = self._conn_params

        self.info_label.setText('불러오는 중...')
        self.schema_table.clear()

        self._schema_thread = QThread(self)
        self._schema_worker = SchemaLoadWorker(
            request_id=request_id,
            dbtype=dbtype,
            host=host,
            port=port,
            user=user,
            pw=pw,
            db=db,
            tables=checked_tables,
        )
        self._schema_worker.moveToThread(self._schema_thread)
        self._schema_thread.started.connect(self._schema_worker.run)
        self._schema_worker.finished.connect(self._on_schema_loaded)
        self._schema_worker.error.connect(self._on_schema_error)

        self._schema_worker.finished.connect(self._schema_thread.quit)
        self._schema_worker.error.connect(self._schema_thread.quit)
        self._schema_thread.finished.connect(self._schema_thread.deleteLater)

        self._schema_thread.start()

    @pyqtSlot(int, str, list, list)
    def _on_schema_loaded(self, request_id: int, info_html: str, headers: list, rows: list):
        # 최신 요청만 반영
        if request_id != self._schema_request_id:
            return

        self.info_label.setText(info_html)

        # 표는 청크로 채워 UI 멈춤 최소화
        self.schema_table.clear()

        if not headers:
            self.schema_table.setRowCount(0)
            self.schema_table.setColumnCount(0)
            self._schema_worker = None
            self._schema_thread = None
            return

        self.schema_table.blockSignals(True)
        self.schema_table.setUpdatesEnabled(False)
        self.schema_table.setColumnCount(len(headers))
        self.schema_table.setHorizontalHeaderLabels([str(h) for h in headers])
        self.schema_table.setRowCount(len(rows))

        self._populate_request_id = request_id
        self._populate_rows = rows
        self._populate_index = 0
        self._populate_timer.start(0)

        self._schema_worker = None
        self._schema_thread = None

    def _populate_next_chunk(self):
        # 최신 요청만 계속 렌더링
        if self._populate_request_id != self._schema_request_id:
            self._populate_timer.stop()
            self.schema_table.blockSignals(False)
            self.schema_table.setUpdatesEnabled(True)
            return

        if not self._populate_rows:
            self._populate_timer.stop()
            self.schema_table.blockSignals(False)
            self.schema_table.setUpdatesEnabled(True)
            return

        start = self._populate_index
        end = min(start + self._populate_chunk_size, len(self._populate_rows))

        col_count = self.schema_table.columnCount()
        for i in range(start, end):
            row = self._populate_rows[i]
            # 행 길이가 헤더와 다를 수 있어 안전하게 처리
            for j in range(col_count):
                val = row[j] if j < len(row) else ''
                self.schema_table.setItem(i, j, QTableWidgetItem(str(val)))

        self._populate_index = end

        if self._populate_index >= len(self._populate_rows):
            self._populate_timer.stop()
            self.schema_table.blockSignals(False)
            self.schema_table.setUpdatesEnabled(True)
            self._populate_rows = []

    @pyqtSlot(int, str)
    def _on_schema_error(self, request_id: int, message: str):
        if request_id != self._schema_request_id:
            return
        self.info_label.setText('오류 발생')
        QMessageBox.critical(self, '오류', message)
        self._schema_worker = None
        self._schema_thread = None

    def export_to_excel(self):
        if not self.conn:
            QMessageBox.warning(self, '경고', '테이블을 선택하세요.')
            return
        tables = []
        table_comments = {}
        for i in range(self.table_list.rowCount()):
            item = self.table_list.item(i, 0)
            if item and item.checkState():
                display = self.table_list.item(i, 1).text()
                tname = display.split(' / ')[0]
                tcomment = display.split(' / ')[1] if ' / ' in display else ''
                tables.append(tname)
                table_comments[tname] = tcomment
        if not tables:
            QMessageBox.warning(self, '경고', '체크된 테이블이 없습니다.')
            return
        save_path, _ = QFileDialog.getSaveFileName(self, '엑셀로 저장', '', 'Excel Files (*.xlsx)')
        if not save_path:
            return

        if not self._conn_params:
            QMessageBox.critical(self, '오류', '연결 정보가 없습니다. 먼저 Connect를 다시 시도하세요.')
            return

        dbtype, host, port, user, pw, db = self._conn_params

        # 백그라운드 내보내기
        self.export_btn.setEnabled(False)
        self.ddl_btn.setEnabled(False)
        self.connect_btn.setEnabled(False)
        self._export_progress = QProgressDialog('엑셀 내보내는 중...', None, 0, len(tables), self)
        self._export_progress.setWindowTitle('Export')
        self._export_progress.setWindowModality(Qt.WindowModal)
        self._export_progress.setAutoClose(False)
        self._export_progress.setAutoReset(False)
        self._export_progress.setValue(0)
        self._export_progress.show()

        self._export_thread = QThread(self)
        self._export_worker = ExcelExportWorker(
            dbtype=dbtype,
            host=host,
            port=port,
            user=user,
            pw=pw,
            db=db,
            tables=tables,
            table_comments=table_comments,
            save_path=save_path,
        )
        self._export_worker.moveToThread(self._export_thread)
        self._export_thread.started.connect(self._export_worker.run)
        self._export_worker.progress.connect(self._on_export_progress)
        self._export_worker.finished.connect(self._on_export_finished)
        self._export_worker.error.connect(self._on_export_error)

        # 스레드 정리
        self._export_worker.finished.connect(self._export_thread.quit)
        self._export_worker.error.connect(self._export_thread.quit)
        self._export_thread.finished.connect(self._export_thread.deleteLater)

        self._export_thread.start()

    @pyqtSlot(int, str)
    def _on_export_progress(self, done: int, message: str):
        if self._export_progress:
            self._export_progress.setLabelText(message)
            self._export_progress.setValue(done)

    @pyqtSlot(str)
    def _on_export_finished(self, save_path: str):
        if self._export_progress:
            self._export_progress.close()
            self._export_progress = None
        self.export_btn.setEnabled(True)
        self.ddl_btn.setEnabled(True)
        self.connect_btn.setEnabled(True)
        QMessageBox.information(self, '완료', '엑셀 저장 완료!')
        self._export_worker = None
        self._export_thread = None

    @pyqtSlot(str)
    def _on_export_error(self, message: str):
        if self._export_progress:
            self._export_progress.close()
            self._export_progress = None
        self.export_btn.setEnabled(True)
        self.ddl_btn.setEnabled(True)
        self.connect_btn.setEnabled(True)
        QMessageBox.critical(self, '에러', message)
        self._export_worker = None
        self._export_thread = None

    def show_ddl(self):
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPlainTextEdit, QPushButton
        if not self.conn:
            QMessageBox.warning(self, '경고', '테이블을 선택하세요.')
            return
        tables = []
        for i in range(self.table_list.rowCount()):
            item = self.table_list.item(i, 0)
            if item and item.checkState():
                tables.append(self.table_list.item(i, 1).text().split(' / ')[0])
        if not tables:
            QMessageBox.warning(self, '경고', '체크된 테이블이 없습니다.')
            return
        ddls = []
        for table in tables:
            ddl = get_table_ddl(self.conn, self.dbtype, table)
            ddls.append(f'-- {table} --\n{ddl}\n')
        dlg = QDialog(self)
        dlg.setWindowTitle('DDL')
        layout = QVBoxLayout()
        text_edit = QPlainTextEdit('\n'.join(ddls))
        text_edit.setReadOnly(False)
        layout.addWidget(text_edit)
        close_btn = QPushButton('닫기')
        close_btn.clicked.connect(dlg.accept)
        layout.addWidget(close_btn)
        dlg.setLayout(layout)
        dlg.resize(700, 500)
        dlg.exec_()

    def save_settings(self):
        data = {
            'dbtype': self.dbtype_combo.currentText(),
            'host': self.host_input.text(),
            'port': self.port_input.text(),
            'user': self.user_input.text(),
            'db': self.db_input.text()
        }
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f)
        # 안내 팝업 제거

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.dbtype_combo.setCurrentText(data.get('dbtype', 'mysql'))
            self.host_input.setText(data.get('host', ''))
            self.port_input.setText(str(data.get('port', '')))
            self.user_input.setText(data.get('user', ''))
            self.db_input.setText(data.get('db', ''))
