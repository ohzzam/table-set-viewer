from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QLabel, QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox, QComboBox, QListWidget, QAbstractItemView)
from PyQt5.QtCore import Qt
import json
import os
import pandas as pd
from db_utils import get_connection, get_tables, get_table_schema, get_table_ddl, export_schema_to_excel

SETTINGS_FILE = 'c:/Python/TABLE_SET/settings.json'

class MainWindow(QWidget):
    def handle_table_cell_click(self, row, col):
        # 테이블명 셀 클릭 시 체크박스 토글
        if col == 1:
            item = self.table_list.item(row, 0)
            if item:
                item.setCheckState(0 if item.checkState() else 2)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("DB Table Viewer (확장판)")
        self.conn = None
        self.dbtype = 'mysql'
        self.init_ui()
        self.load_settings()

    def init_ui(self):
        layout = QVBoxLayout()
        # DB 입력 폼
        form_layout = QHBoxLayout()
        self.dbtype_combo = QComboBox(); self.dbtype_combo.addItems(['mysql', 'postgresql', 'cubrid'])
        self.host_input = QLineEdit(); self.host_input.setPlaceholderText('Host')
        self.port_input = QLineEdit(); self.port_input.setPlaceholderText('Port')
        self.user_input = QLineEdit(); self.user_input.setPlaceholderText('User')
        self.pw_input = QLineEdit(); self.pw_input.setPlaceholderText('Password'); self.pw_input.setEchoMode(QLineEdit.Password)
        self.db_input = QLineEdit(); self.db_input.setPlaceholderText('Database')
        self.connect_btn = QPushButton('Connect'); self.connect_btn.clicked.connect(self.connect_db)
        self.pw_input.returnPressed.connect(self.connect_db)
        form_layout.addWidget(QLabel('DB Type'))
        form_layout.addWidget(self.dbtype_combo)
        form_layout.addWidget(self.host_input)
        form_layout.addWidget(self.port_input)
        form_layout.addWidget(self.user_input)
        form_layout.addWidget(self.pw_input)
        form_layout.addWidget(self.db_input)
        form_layout.addWidget(self.connect_btn)
        layout.addLayout(form_layout)

        # 테이블 목록 (체크박스/전체선택)
        from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QCheckBox, QHeaderView
        self.select_all_checkbox = QCheckBox('전체선택')
        self.select_all_checkbox.stateChanged.connect(self.select_all_tables)
        layout.addWidget(self.select_all_checkbox)
        self.table_list = QTableWidget()
        self.table_list.setColumnCount(2)
        self.table_list.setHorizontalHeaderLabels(['선택', '테이블 (이름 / 코멘트)'])
        self.table_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table_list.cellChanged.connect(self.show_table_schema)
        self.table_list.cellClicked.connect(self.handle_table_cell_click)
        layout.addWidget(self.table_list)

        # 테이블 정보(이름, 코멘트, PK, FK, 인덱스) 표시 (스크롤 가능)
        from PyQt5.QtWidgets import QScrollArea
        self.info_label = QLabel('')
        self.info_label.setTextInteractionFlags(self.info_label.textInteractionFlags() | Qt.TextSelectableByMouse)
        self.info_label.setWordWrap(True)
        self.info_scroll = QScrollArea()
        self.info_scroll.setWidgetResizable(True)
        self.info_scroll.setWidget(self.info_label)
        self.info_scroll.setMinimumHeight(180)
        layout.addWidget(self.info_scroll)

        # 테이블 구조 표시
        self.schema_table = QTableWidget()
        layout.addWidget(QLabel('테이블 구조'))
        layout.addWidget(self.schema_table)
        self.schema_table.verticalHeader().setVisible(False)

        # 버튼들
        btn_layout = QHBoxLayout()
        self.export_btn = QPushButton('엑셀로 내보내기'); self.export_btn.clicked.connect(self.export_to_excel)
        self.ddl_btn = QPushButton('DDL 생성'); self.ddl_btn.clicked.connect(self.show_ddl)
        self.save_btn = QPushButton('접속정보 저장'); self.save_btn.clicked.connect(self.save_settings)
        btn_layout.addWidget(self.export_btn)
        btn_layout.addWidget(self.ddl_btn)
        btn_layout.addWidget(self.save_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def connect_db(self):
        dbtype = self.dbtype_combo.currentText()
        host = self.host_input.text()
        port = int(self.port_input.text())
        user = self.user_input.text()
        pw = self.pw_input.text()
        db = self.db_input.text()
        try:
            self.conn = get_connection(dbtype, host, port, user, pw, db)
            self.dbtype = dbtype
            tables = get_tables(self.conn, dbtype)
            self.table_list.setRowCount(len(tables))
            for i, (name, comment) in enumerate(tables):
                display = f"{name} / {comment}" if comment else name
                chk_item = QTableWidgetItem()
                chk_item.setCheckState(0)
                self.table_list.setItem(i, 0, chk_item)
                self.table_list.setItem(i, 1, QTableWidgetItem(display))
            self.save_settings()  # 연결 성공 시 자동 저장
            QMessageBox.information(self, '성공', 'DB 연결 및 테이블 목록 불러오기 성공!')
        except Exception as e:
            QMessageBox.critical(self, '오류', str(e))
    def select_all_tables(self, state):
        for i in range(self.table_list.rowCount()):
            item = self.table_list.item(i, 0)
            if item:
                item.setCheckState(2 if state else 0)

    def show_table_schema(self):
        # 체크된 모든 테이블 구조를 아래에 모두 표시
        if not self.conn:
            return
        checked_tables = []
        for i in range(self.table_list.rowCount()):
            item = self.table_list.item(i, 0)
            if item and item.checkState():
                checked_tables.append(self.table_list.item(i, 1).text())
        if not checked_tables:
            self.info_label.setText('')
            self.schema_table.clear()
            return
        # 여러 테이블 정보/구조를 HTML로 합쳐 info_label에, 구조는 QTableWidget에 모두 표시
        info_html = ''
        all_rows = []
        all_headers = None
        for table_display in checked_tables:
            table = table_display.split(' / ')[0]
            schema = get_table_schema(self.conn, self.dbtype, table)
            columns = schema['columns']
            indexes = schema['indexes']
            pk = schema['primary_key']
            fk = schema['foreign_keys']
            comment = schema['table_comment']
            # 정보 표시
            info = f"<b>Table:</b> {table}<br>"
            info += f"<b>Comment:</b> {comment}<br>"
            info += f"<b>Primary Key:</b> {', '.join(pk) if pk else '-'}<br>"
            if fk:
                fk_str = '<br>'.join([f"{c} → {t}({col})" for c, t, col in fk])
            else:
                fk_str = '-'
            info += f"<b>Foreign Key:</b> {fk_str}<br>"
            if indexes:
                idx_str = '<br>'.join([f"<b>{t}</b> <span style='color:blue'>{iname}</span> ({icols})" for t, iname, icols in indexes])
            else:
                idx_str = '-'
            info += f"<b>Index info:</b> {idx_str}"
            info_html += info + '<hr>'
            # 구조 표시 (No 컬럼만, 테이블별 1부터 시작, 구분선)
            if self.dbtype == 'mysql':
                headers = ['No', 'Table', 'Field', 'Type', 'Null', 'Key', 'Default', 'Extra', 'Comment']
                if all_headers is None:
                    all_headers = headers
                for idx, col in enumerate(columns, 1):
                    null_val = 'NN' if col[3] == 'NO' else ''
                    row = [idx, table, col[0], col[1], null_val, col[4], col[5], col[6], col[8]]
                    all_rows.append(row)
                all_rows.append([''] * len(headers))
            elif self.dbtype == 'postgresql':
                headers = ['No', 'Table', 'column_name', 'data_type', 'is_nullable', 'column_default']
                if all_headers is None:
                    all_headers = headers
                for idx, col in enumerate(columns, 1):
                    row = [idx, table] + list(col)
                    all_rows.append(row)
                all_rows.append([''] * len(headers))
        self.info_label.setText(info_html)
        # 구조 테이블에 모두 표시
        self.schema_table.clear()
        if all_headers:
            # 마지막 빈 행 제거
            if all_rows and all_rows[-1] == [''] * len(all_headers):
                all_rows.pop()
            self.schema_table.setColumnCount(len(all_headers))
            self.schema_table.setHorizontalHeaderLabels(all_headers)
            self.schema_table.setRowCount(len(all_rows))
            for i, row in enumerate(all_rows):
                for j, val in enumerate(row):
                    self.schema_table.setItem(i, j, QTableWidgetItem(str(val)))

    def export_to_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            if not self.conn:
                QMessageBox.warning(self, '경고', '테이블을 선택하세요.')
                return
            tables = []
            table_comments = {}
            for i in range(self.table_list.rowCount()):
                item = self.table_list.item(i, 0)
                if item and item.checkState():
                    display = self.table_list.item(i, 1).text()
                    tname = display.split(' / ')[0]
                    tcomment = display.split(' / ')[1] if ' / ' in display else ''
                    tables.append(tname)
                    table_comments[tname] = tcomment
            if not tables:
                QMessageBox.warning(self, '경고', '체크된 테이블이 없습니다.')
                return
            save_path, _ = QFileDialog.getSaveFileName(self, '엑셀로 저장', '', 'Excel Files (*.xlsx)')
            if not save_path:
                return
            wb = openpyxl.Workbook()
            # 1. 테이블 목록 시트
            ws1 = wb.active
            ws1.title = '테이블 목록'
            ws1.append(['NN', 'Table Name', 'Description'])
            for idx, t in enumerate(tables, 1):
                ws1.append([idx, t, table_comments.get(t, '')])
            # 스타일 적용 (헤더 회색, 굵게, 가운데)
            header_fill = PatternFill('solid', fgColor='B7B7B7')
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in range(1, 4):
                ws1.column_dimensions[get_column_letter(col)].width = 20
            # 2. 테이블 정의서 시트
            ws2 = wb.create_sheet('테이블 정의서')
            row_cursor = 1
            table_blocks = []  # (start_row, end_row) for each table
            for t in tables:
                block_start = row_cursor
                schema = get_table_schema(self.conn, self.dbtype, t)
                columns = schema['columns']
                pk = schema['primary_key']
                fk = schema['foreign_keys']
                indexes = schema['indexes']
                comment = schema['table_comment']
                # 테이블 정보 블록
                ws2.merge_cells(start_row=row_cursor, start_column=2, end_row=row_cursor, end_column=8)
                ws2.cell(row=row_cursor, column=2, value='테이블 정의서').fill = header_fill
                ws2.cell(row=row_cursor, column=2).font = Font(bold=True)
                ws2.cell(row=row_cursor, column=2).alignment = Alignment(horizontal='center')
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Table ID')
                ws2.cell(row=row_cursor, column=3, value=t)
                ws2.cell(row=row_cursor, column=4, value='Table Name')
                ws2.cell(row=row_cursor, column=5, value=t)
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Description')
                ws2.cell(row=row_cursor, column=3, value=comment)
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Primary Key')
                ws2.cell(row=row_cursor, column=3, value=','.join(pk) if pk else '')
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Foreign Key')
                fk_str = ', '.join([f"{c}->{t2}({col})" for c, t2, col in fk]) if fk else ''
                ws2.cell(row=row_cursor, column=3, value=fk_str)
                row_cursor += 1
                ws2.cell(row=row_cursor, column=2, value='Index info #1')
                row_cursor += 1
                # 인덱스 정보 표
                if indexes:
                    idx_headers = ['Index Name', 'Columns', 'Unique']
                    for i, h in enumerate(idx_headers, 2):
                        ws2.cell(row=row_cursor, column=i, value=h)
                        ws2.cell(row=row_cursor, column=i).fill = header_fill
                        ws2.cell(row=row_cursor, column=i).font = Font(bold=True)
                        ws2.cell(row=row_cursor, column=i).alignment = Alignment(horizontal='center')
                    row_cursor += 1
                    for idx in indexes:
                        iname = idx[1]
                        icols = idx[2]
                        unique = 'Y' if len(idx) > 3 and idx[3] else ''
                        ws2.cell(row=row_cursor, column=2, value=iname)
                        ws2.cell(row=row_cursor, column=3, value=icols)
                        ws2.cell(row=row_cursor, column=4, value=unique)
                        row_cursor += 1
                else:
                    row_cursor += 1
                # 컬럼 헤더 (NN~Default까지 7개 컬럼, 모두 회색)
                col_headers = ['NN', 'Physical Name', 'Logical Name', 'Data Type', 'Null', 'Key', 'Default']
                for i, h in enumerate(col_headers, 2):
                    ws2.cell(row=row_cursor, column=i, value=h)
                    ws2.cell(row=row_cursor, column=i).fill = header_fill
                    ws2.cell(row=row_cursor, column=i).font = Font(bold=True)
                    ws2.cell(row=row_cursor, column=i).alignment = Alignment(horizontal='center')
                row_cursor += 1
                # 컬럼 정보
                for idx, col in enumerate(columns, 1):
                    null_val = 'NN' if col[3] == 'NO' else ''
                    key_val = col[4]
                    ws2.cell(row=row_cursor, column=2, value=idx)
                    ws2.cell(row=row_cursor, column=3, value=col[0])
                    ws2.cell(row=row_cursor, column=4, value=col[8] if len(col) > 8 else '')
                    ws2.cell(row=row_cursor, column=5, value=col[1])
                    ws2.cell(row=row_cursor, column=6, value=null_val)
                    ws2.cell(row=row_cursor, column=7, value=key_val)
                    ws2.cell(row=row_cursor, column=8, value=col[5])
                    row_cursor += 1
                block_end = row_cursor - 1
                table_blocks.append((block_start, block_end))
                row_cursor += 2  # 테이블별 구분을 위해 빈 줄 2개
            # 스타일: 전체 표 테두리(테이블 정의서 시트는 각 테이블 블록에만 적용)
            thin = Side(border_style='thin', color='000000')
            for start, end in table_blocks:
                for row in ws2.iter_rows(min_row=start, max_row=end, min_col=2, max_col=8):
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            # 테이블 목록 시트는 전체 테두리 유지
            for row in ws1.iter_rows():
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            # 자동 높이/너비
            for ws in [ws1, ws2]:
                for col in ws.columns:
                    maxlen = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                maxlen = max(maxlen, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max(12, min(maxlen+2, 40))
            wb.save(save_path)
            QMessageBox.information(self, '완료', '엑셀 저장 완료!')
        except Exception as e:
            QMessageBox.critical(self, '에러', f'엑셀 저장 중 오류 발생: {e}')

    def show_ddl(self):
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPlainTextEdit, QPushButton
        if not self.conn:
            QMessageBox.warning(self, '경고', '테이블을 선택하세요.')
            return
        tables = []
        for i in range(self.table_list.rowCount()):
            item = self.table_list.item(i, 0)
            if item and item.checkState():
                tables.append(self.table_list.item(i, 1).text().split(' / ')[0])
        if not tables:
            QMessageBox.warning(self, '경고', '체크된 테이블이 없습니다.')
            return
        ddls = []
        for table in tables:
            ddl = get_table_ddl(self.conn, self.dbtype, table)
            ddls.append(f'-- {table} --\n{ddl}\n')
        dlg = QDialog(self)
        dlg.setWindowTitle('DDL')
        layout = QVBoxLayout()
        text_edit = QPlainTextEdit('\n'.join(ddls))
        text_edit.setReadOnly(False)
        layout.addWidget(text_edit)
        close_btn = QPushButton('닫기')
        close_btn.clicked.connect(dlg.accept)
        layout.addWidget(close_btn)
        dlg.setLayout(layout)
        dlg.resize(700, 500)
        dlg.exec_()

    def save_settings(self):
        data = {
            'dbtype': self.dbtype_combo.currentText(),
            'host': self.host_input.text(),
            'port': self.port_input.text(),
            'user': self.user_input.text(),
            'db': self.db_input.text()
        }
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f)
        # 안내 팝업 제거

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.dbtype_combo.setCurrentText(data.get('dbtype', 'mysql'))
            self.host_input.setText(data.get('host', ''))
            self.port_input.setText(str(data.get('port', '')))
            self.user_input.setText(data.get('user', ''))
            self.db_input.setText(data.get('db', ''))
